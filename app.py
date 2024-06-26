# This is a comment
## This is comment
### This is comment
#### This is comment
##### This is comment
"""Multi line"""
'''multi line'''

"""
print("Wanglem")
print('o----')
print(' ||||')
print('*' * 10)

#Python variables
price = 10
print(price)
price = 20
print(price)
rating = 4.9
name = 'Mosh'
is_published = True

# exercise
full_name = 'John Smith'
age = 20
is_new = True

# exercise input()
name = input('What is your name? ')
favorite_color = input('What is your favorite color? ')
print(name + ' likes ' + favorite_color)

# exercise int and str
birth_year = input('Birth year: ')
print(type(birth_year))
age = 2019 - int(birth_year)
print(type(age))
print(age)

# Exercise float and str/int
weight_lbs = input(' Weight (lbs): ')
weight_kg = int(weight_lbs) * 0.45
print(weight_kg)

# Strings/str
# course = "Python's Course for Beginners"
# course = 'Python for "Beginners"'
course = '''
Hi John,

Here is our first email to you.

Thankyou,
The support team

'''
print(course)

# Indexing string
course = 'Python for Beginners'
print(course[0])
print(course[-1])
print(course[0:3])
print(course[0:])
print(course[:5])

another = course[:]
print(another)

# Indexing upto -1 from 1
name = 'Jennifer'
print(name[1:-1])

# To achieve John [Smith] is a coder
first = 'John'
last = 'Smith'
message = first + ' [' + last + '] is a coder'
# Using formatted string for easier understanding of concatenation
# Prefix by 'f' to define formatted string
msg = f'{first} [{last}] is a coder'
print(message)
print(msg)

# String Methods
course = 'Python for Beginners'
print(len(course))
print(course.upper())
print(course.lower())
print(course)
print(course.find('P'))  # It's case sensitive,and it return the first value in index,not entire
print(course.replace('Beginners', 'Absolute Beginners'))  # This is also case sensitive
print('Python' in course)  # If exists.Also case sensitive

# Arithmetic operations
print(10 / 3)  # Gets floating point
print(10 // 3)  # Gets integer
print(10 % 3)
print(10 ** 3)

# Augmented assignment operator
x = 10
x = x + 3
x += 3
print(x)

# Precedent orders Brackets, Exponentiation 2 ** 3, Multiplication and division from right to left, Addition and subtraction from right to left
x = 10 + 3 * 2
y = 21 - 13 + 10
print(x)
print(y)

# Maths functions to work with numbers
x = 2.9
print(round(x))
# Absolute function
y = -2.9
print(abs(y))  # Returns positive 2.9
print(abs(-2.9))  # Returns positive number

# Math module has many functions to be use
import math

print(math.ceil(2.9))  # Just rounds off
print(math.floor(2.9))  # Rounds off to low

# if elif else statements
is_hot = True
is_cold = True

if is_hot:
    print("It's a hot day")
    print("Drink plenty of water")
elif is_cold:
    print("It's a cold day")
    print("Wear warm clothes")
else:
    print("It's a lovely day")
print("Enjoy your day")

# if elif else Exercise

price = 1000000
good_credit = True

if good_credit:
    down_payment = 0.1 * price
else:
    down_payment = 0.2 * price
print(f"Down payment: ${down_payment}")

# Logical and/or/not operators for dealing with multiple conditions in one statements
has_high_income = True
has_good_credit = False
has_criminal_record = True

if has_high_income and has_good_credit:
    print("Eligible for Loan")

if has_high_income or has_good_credit:
    print("Eligible for loan")

if has_high_income and not has_criminal_record:
    print("Eligible for loan")
    
# Comparison Operators
temperature = 30

if temperature > 30:
    print("It's a hot day")
else:
    print("It's not a hot day")

# Validation Exercise
name = input("Enter your name: ")

if len(name) < 3:
    print("Name too short")
elif len(name) > 15:
    print("Name too long")
else:
    print("Good name")
 
# Conversion Exercise
weight = int(input('Weight: '))
unit = input('(L)bs or (K)g: ')

if unit.upper() == "L":
    converted = weight * 0.5
    print(f"You are {converted} Kilos")
else:
    converted = weight / 0.5
    print(f"You are {converted} Pounds")

# While loop
i = 1
while i <= 5:
    print(i)
    i = i + 1
print("Done")
while i <= 5:
    print('*' * i)
    i = i + 1
print("Done")

# Guess game
secret_number = 9
guess_count = 0
guess_limit = 3
while guess_count < guess_limit:
    guess = int(input('Guess: '))
    guess_count += 1
    if guess == secret_number:
        print('You Won')
        break
else:
    print('You failed')

# Car game Exercise
command = ""
started = False

# while command != "quit":
while True:

    command = input("> ").lower()
    if command == "start":
        if started:
            print("Car is already started")
        else:
            started = True
            print("Car started...")
    elif command == "stop":
        if not started:
            print("Car is already stopped")
        else:
            started = False
            print("Car stopped...")
    elif command == "help":
        print('''
start - to start the car
stop - to stop the car
quit - to quit
        ''')
    elif command == "quit":
        break
    else:
        print("Sorry, I don't understand that")

"""
"""
# For loops
for item in 'Python':
    print(item)
    
for item in ['Mosh', 'John', 'Sarah']:
    print(item)
    
for item in range(10): # range(5, 10) this means 5 to 10. Also range(5, 10, 2) this means from 5 to 10,but 2 steps forward,it will skip the numbers according to what parameter takes
    print(item)

# for loop exercise,sum up the total price in an array
price = [10, 20, 30]
total = 0

for num in price:
    total += num
print(f"Total price is: {total}")

# Nested loop,creating x and y coordinate
for x in range(4):
    for y in range(3):
        print(f'({x}, {y})')

# For loop exercise
numbers = [5, 2, 5, 2, 2]

for x_count in numbers:
    # print('x' * x) This is shortcut without nested loop
    output = ''
    for count in range(x_count):
        output += 'x'
    print(output)
      
# Lists methods
names = ['John', 'Bob', 'Mosh', 'Sarah', 'Mary']
print(names)
print(names[0])
print(names[-1])
print(names[2:5])
print(names[3:])

# List exercise,find the largest number
# my way
numbers = [3, 6, 10, 2, 8, 4]
largest = 0
for x in numbers:
    if x > largest:
        largest = x
    else:
        largest = largest
print(largest)

# Mosh Way
numbers = [3, 6, 10, 2, 8, 4]
max = numbers[0]
for number in numbers:
    if number > max:
        max = number
print(max)

# 2D list,mostly used in DS and ML
matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
print(matrix[0][1])
matrix[0][1] = 20 # Modification
print(matrix[0][1])
for row in matrix:
    for item in row:
        print(item)

# List methods/operation/functions
numbers = [5, 2, 4, 9, 8]
numbers.append(20) # Adding number or value
numbers.insert(0, 10) # Insert allows to add number or value to the index we want
numbers.remove(5) # Only the matching value
numbers.clear() # Clears all
numbers.pop() # Remove the number or value at the last index
print(numbers.index(4)) # Check if it exist in list,and returns in which index it exists if
print(50 in numbers) # To get Boolean value whether number exists
print(numbers.count(5)) # Counts numbers of same value,or how many same number exist
numbers.sort() # For sorting,it returns none,none is also a vlue in python
numbers.reverse() # For sorting in descending order
numbers2 = numbers.copy() # Create another copy of it,make changes will not impact original

# Exercise, Removing duplicates
# My way
numbers = [2, 2, 4, 6, 3, 4, 6, 1]
uniques = []
for num in numbers:
    if num in uniques:
        uniques = uniques
    else:
        uniques.append(num)
print(uniques)
# Mosh way
numbers = [2, 2, 4, 6, 3, 4, 6, 1]
uniques = []
for num in numbers:
    if num not in uniques:
        uniques.append(num)
print(uniques)

# Tuples
numbers = (1, 2, 3) # tuples uses () instead of []
# Tuple has only 2 methods, i.e. to count and find the index of a value
print(numbers[0])

# Unpacking
coordinates = (1, 2, 3)
"""
"""
x = coordinates[0]
y = coordinates[1]
z = coordinates[2]

# Tuples can easily make it easier than above
x, y, z = coordinates # This can also be used with list,not just tuples
print(x)

# Dictionaries. Used for storing key value pairs
customer = {
    "name": "John",
    "age": 30,
    "is_verified": True
}
print(customer["age"]) # Case sensitive
print(customer.get("age", "Jan 1 1980"))  # This can also take default value
customer["name"] = "Jack Smith"
print(customer)
customer["birthdate"] = " Jan 1 1980" # Adding another key value pair
print(customer)

# Dictionaries Exercise,translate numerical ph number to word
phone = input("Phone: ")
numbers = {
    "0": "zero",
    "1": "one",
    "2": "two",
    "3": "three",
    "4": "four"
}
output = ""
for ch in phone:
    output += numbers.get(ch, "!") + " "
print(output)

# Emoji converter
message = input(">")
words = message.split(' ')
emojis = {
    ":(": "😔",
    ":)": "😊"
}
output = ""
for word in words:
    output += emojis.get(word, word) + " "
print(output)

# Functions
def greet_user():
    print("Hi there!")
    print("Welcome aboard")


print("Start")
greet_user()

# Parameters
apple = "Apple"
def greet_user(name, user):
    name = "John"
    its = user
    print(F'Hi {name}')
    print(f"It's {user}")


greet_user("John", apple)

# Keyword Arguments. Keyword position doesn't matter,in which index in parameter value is given
apple = "Apple"
name = "John Smith"
def greet_user(name, user):
    name = name
    its = user
    print(F'Hi {name}')
    print(f"It's {user}")


greet_user(name, apple)
greet_user(apple, name)

# Return Statement. By default function returns None.
def return_square(square):
    return square * square


return_square(4)
result = return_square(5)
print(result)
print(return_square(4))

# Exercise.Creating a reusable function
# My way
def return_emoji():
    emo = input(">")
    result = ""
    words = emo.split(' ')
    emojis = {
        ":(": "😔",
        ":)": "😊"
    }
    for emos in emojis:
        result = emojis.get(emos)
    return result

return_emoji()
print(return_emoji())

# Mosh way
def emoji_converter(message):
    words = message.split(' ')
    emojis = {
        ":(": "😔",
        ":)": "😊"
    }
    output = " "
    for word in words:
        output += emojis.get(word, word)
    return output


message = input(">")
result = emoji_converter(message)
print(result)

# Exceptions or Handling error
try:
    age = int(input('Age: '))
    income = 20000
    rist = income / age
    print(age)
except ZeroDivisionError:
    print("Age cannot be 0")
except ValueError:
    print("Invalid Value")
    
# Classes
class Point:
    def move(self):
        print("move")
    def draw(self):
        print("draw")


point1 = Point()
point1.draw()
point1.move()
point1.x = 10
point1.y = 20
print(point1.x)
print(point1.y)

# Constructor
class Point:
    def __init__(self, x, y): # This is constructor
        self.x = x
        self.y = y
    def move(self):
        print("move")
    def draw(self):
        print("draw")


point = Point(10, 20)
print(point.x)
point.x = 11
print(point.x)

# Exercise
class Person:
    def __init__(self, name):
        self.name = name
    def talk(self):
        print(f"Hi, I am {self.name}")


john = Person("John Smith")
print(john.name)
john.talk()

# Inheritance. DRY! Don't Repeat Yourself
class Mammal:
    def walk(self):
        print("Walk")


class Dog(Mammal): # Inheriting all the methods from Mammal class
    def bark(self):
        print("Bark")


class Cat(Mammal):
    pass # pass statement used incase there's nothing to pass

dog1 = Dog()
dog1.walk()
dog1.bark()

# Modules
import converter # direct from converter.py without any export function like js
from converter import lbs_to_kg # To import only specific function

lbs_to_kg(100) # it can be used like other function because it was imported specifically

print(converter.kg_to_lbs(70))

# Exercise. Make a function in other module and call the function here to find the largest number in the array given below
# My way
numbers = [10, 3, 13, 9, 8]

from utils import largest_num

largest_num(numbers)

# Mosh way
from utils import find_max

print(find_max(numbers))

# Packages
from ecommerce.shipping import calc_shipping # Another way of import for direct use of function like below. To import more use comma ","
import ecommerce.shipping # Folder/file
ecommerce.shipping.calc_shipping() # Folder/file
calc_shipping() # Direct use of function

# Generate random number by built-in module
import random

for i in range(3):
    print(random.random()) # 0 t0 1
    print(random.randint(10, 20)) # 10 to 20

# Picking leader randomly
members = ['John', 'Mary', 'Bob', 'Mosh']
leader = random.choice(members)
print(leader)

# Exercise.Roll a dice using random func,class and tuple
import random

class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return (first, second) # () means tuple,or just return first, second without bracket,it'll automatically return as tuple


dice = Dice()
print(dice.roll())

# Files and Directories
from pathlib import Path # Path is a class,because it has 'P' in upper case
# Absolute path - c:\Program Files\Microsoft - /usr/local/bin
# Relative path - It's path to the package or module in project directory
path = Path("ecommerce") # working with relative path
print(path.exists()) # checking if path exists,result is boolean
path_that_donExist = Path("emails")
print(path_that_donExist.exists())
print(path_that_donExist.mkdir()) # making directory that don't exist,prints 'none' but the directory is found in project directory
print(path_that_donExist.rmdir()) # Now to remove

path = Path()
for file in path.glob('*.py'): # For all files and folders ('*')
    print(file) # Get all the .py files in the directory
    
'''
"""
# Pypi and Pip (Huh?)
# pip install openpyxl,to install package to work with xl

# Working with xl files (Create excel file for this tutorial)
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('mosh.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)
#print(cell.value) # Checking value
#print(sheet.max_row) # Checking number of rows

for row in range(2, sheet.max_row + 1): # start from 2 to ignore heading or column name
    # print(row)
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4) # Creating new cell,instead of overwriting that cell with wrong price
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('mosh3.xlsx') # Run and check this project directory,new file is created